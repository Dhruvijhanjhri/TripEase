import csv

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse

from bookings.models import Booking
from hotels.models import HotelBooking
from packages.models import PackageBooking
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from datetime import datetime
from django.db.models import Sum
from payments.models import Payment
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.colors import HexColor

def get_export_data():
    """
    Returns a unified list of all bookings
    for CSV, Excel and PDF exports.
    """

    rows = []

    # -----------------------------
    # Flight Bookings
    # -----------------------------
    flights = (
        Booking.objects
        .select_related(
            "user",
            "flight__source",
            "flight__destination",
            "payment"
        )
    )

    for booking in flights:

        payment = getattr(booking, "payment", None)

        route = "N/A"

        if booking.flight:
            route = (
                f"{booking.flight.source.code}"
                f" → "
                f"{booking.flight.destination.code}"
            )

        rows.append([
            booking.booking_reference,
            "Flight",
            booking.user.get_full_name() or booking.user.username,
            route,
            booking.booking_status.title(),
            payment.payment_status.title() if payment else "-",
            payment.get_payment_method_display() if payment else "-",
            float(payment.amount) if payment else "",
            booking.created_at.strftime("%d-%m-%Y %H:%M"),
        ])

    # -----------------------------
    # Hotel Bookings
    # -----------------------------
    hotels = (
        HotelBooking.objects
        .select_related(
            "user",
            "hotel",
            "payment"
        )
    )

    for booking in hotels:

        payment = getattr(booking, "payment", None)

        rows.append([
            booking.booking_reference,
            "Hotel",
            booking.user.get_full_name() or booking.user.username,
            booking.hotel.name if booking.hotel else "N/A",
            booking.booking_status.title(),
            payment.payment_status.title() if payment else "-",
            payment.get_payment_method_display() if payment else "-",
            float(payment.amount) if payment else "",
            booking.created_at.strftime("%d-%m-%Y %H:%M"),
        ])

    # -----------------------------
    # Package Bookings
    # -----------------------------
    packages = (
        PackageBooking.objects
        .select_related(
            "user",
            "package",
            "payment"
        )
    )

    for booking in packages:

        payment = getattr(booking, "payment", None)

        rows.append([
            booking.booking_reference,
            "Package",
            booking.user.get_full_name() or booking.user.username,
            booking.package.name if booking.package else "N/A",
            booking.booking_status.title(),
            payment.payment_status.title() if payment else "-",
            payment.get_payment_method_display() if payment else "-",
            float(payment.amount) if payment else "",
            booking.created_at.strftime("%d-%m-%Y %H:%M"),
        ])

    return rows

@staff_member_required
def export_csv(request):

    response = HttpResponse(
        content_type="text/csv; charset=utf-8"
    )

    response["Content-Disposition"] = (
        'attachment; filename="tripease_dashboard_report.csv"'
    )

    response.write("\ufeff")

    writer = csv.writer(response)

    writer.writerow([
        "Booking Reference",
        "Booking Type",
        "Customer",
        "Details",
        "Booking Status",
        "Payment Status",
        "Payment Method",
        "Amount (₹)",
        "Booking Date",
    ])

    rows = get_export_data()

    for row in rows:
        writer.writerow(row)

    return response

@staff_member_required
def export_excel(request):
    """
    Export dashboard data as a professional Excel report.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "TripEase Dashboard Report"

    headers = [
        "Booking Reference",
        "Booking Type",
        "Customer",
        "Details",
        "Booking Status",
        "Payment Status",
        "Payment Method",
        "Amount (₹)",
        "Booking Date",
    ]

    # Header Style
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    center = Alignment(horizontal="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col_num)

        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Write booking rows
    rows = get_export_data()

    for row in rows:
        ws.append(row)

    # Auto-size columns
    for column in ws.columns:

        max_length = 0

        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 4, 40)

        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="tripease_dashboard_report.xlsx"'
    )

    wb.save(response)

    return response


def format_currency(amount):
    try:
        return f"Rs. {float(amount):,.2f}"
    except:
        return "Rs. 0.00"


def add_page_number(canvas, doc):
    canvas.saveState()

    canvas.setFont("Helvetica", 9)

    canvas.drawString(
        40,
        25,
        "Generated by TripEase Admin Dashboard"
    )

    canvas.drawRightString(
        550,
        25,
        f"Page {doc.page}"
    )

    canvas.restoreState()

@staff_member_required
def export_pdf(request):

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        'attachment; filename="tripease_dashboard_report.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        rightMargin=25,
        leftMargin=25,
        topMargin=30,
        bottomMargin=45,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    title_style.alignment = TA_CENTER
    title_style.fontSize = 22
    title_style.spaceAfter = 20

    heading_style = styles["Heading2"]
    heading_style.spaceAfter = 10

    normal_style = styles["BodyText"]

    elements = []

    # ====================================================
    # TITLE
    # ====================================================

    elements.append(
        Paragraph(
            "<b>TripEase</b>",
            title_style
        )
    )

    elements.append(
        Paragraph(
            "<b>ADMIN ANALYTICS REPORT</b>",
            styles["Heading1"]
        )
    )

    elements.append(Spacer(1, 8))

    generated = datetime.now().strftime("%d %b %Y %I:%M %p")

    elements.append(
        Paragraph(
            f"<b>Generated On:</b> {generated}",
            normal_style
        )
    )

    elements.append(Spacer(1, 20))

    # ====================================================
    # SUMMARY
    # ====================================================

    total_flights = Booking.objects.count()
    total_hotels = HotelBooking.objects.count()
    total_packages = PackageBooking.objects.count()

    total_bookings = (
        total_flights +
        total_hotels +
        total_packages
    )

    successful_payments = Payment.objects.filter(
        payment_status="success"
    ).count()

    total_revenue = (
        Payment.objects.filter(
            payment_status="success"
        ).aggregate(
            total=Sum("amount")
        )["total"] or 0
    )

    elements.append(
        Paragraph(
            "<b>EXECUTIVE SUMMARY</b>",
            heading_style
        )
    )

    summary_data = [

        ["Metric", "Value"],

        ["Total Revenue", format_currency(total_revenue)],

        ["Total Bookings", total_bookings],

        ["Flight Bookings", total_flights],

        ["Hotel Bookings", total_hotels],

        ["Package Bookings", total_packages],

        ["Successful Payments", successful_payments],

    ]

    summary_table = Table(
        summary_data,
        colWidths=[220,180]
    )

    summary_table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),HexColor("#1F4E78")),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BACKGROUND",(0,1),(-1,-1),colors.beige),

            ("GRID",(0,0),(-1,-1),0.4,colors.grey),

            ("BOTTOMPADDING",(0,0),(-1,0),8),

            ("ALIGN",(1,1),(-1,-1),"RIGHT"),

        ])

    )

    elements.append(summary_table)

    elements.append(Spacer(1,25))

    # ====================================================
    # BOOKINGS TABLE
    # ====================================================

    elements.append(
        Paragraph(
            "<b>BOOKING DETAILS</b>",
            heading_style
        )
    )

    rows = get_export_data()

    table_data = [[

        "Ref",

        "Type",

        "Customer",

        "Details",

        "Status",

        "Payment",

        "Method",

        "Amount",

        "Date",

    ]]

    table_data.extend(rows)

    table = Table(

        table_data,

        repeatRows=1,

        colWidths=[

            60,

            45,

            70,

            120,

            55,

            55,

            65,

            60,

            75,

        ],

    )

    table.setStyle(

        TableStyle([

            ("BACKGROUND",(0,0),(-1,0),HexColor("#1F4E78")),

            ("TEXTCOLOR",(0,0),(-1,0),colors.white),

            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),

            ("BOTTOMPADDING",(0,0),(-1,0),8),

            ("GRID",(0,0),(-1,-1),0.25,colors.grey),

            ("FONTSIZE",(0,0),(-1,-1),8),

            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),

            ("ROWBACKGROUNDS",
             (0,1),
             (-1,-1),
             [colors.white, colors.beige]),

            ("ALIGN",(1,1),(6,-1),"CENTER"),

            ("ALIGN",(7,1),(7,-1),"RIGHT"),

        ])

    )

    elements.append(table)

    doc.build(
        elements,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    return response